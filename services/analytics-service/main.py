"""
Analytics Service - FastAPI + Kafka Consumer
Consumes purchase/payment events, generates XLSX/CSV summaries, saves to S3/MinIO.
"""

import asyncio
import csv
import io
import json
import logging
import os
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from typing import Any

import boto3
import openpyxl
import pandas as pd
from aiokafka import AIOKafkaConsumer
from botocore.client import Config
from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse

# ─── Logging ──────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
)
logger = logging.getLogger("analytics")

# ─── Configuration ────────────────────────────────────────────────────────────

KAFKA_BROKERS = os.getenv("KAFKA_BROKERS", "localhost:9092")
KAFKA_GROUP_ID = os.getenv("KAFKA_GROUP_ID", "analytics-group")
S3_ENDPOINT = os.getenv("S3_ENDPOINT", "http://localhost:9000")
S3_ACCESS_KEY = os.getenv("S3_ACCESS_KEY", "minio_admin")
S3_SECRET_KEY = os.getenv("S3_SECRET_KEY", "minio_password")
S3_BUCKET = os.getenv("S3_BUCKET", "analytics-reports")
PORT = int(os.getenv("PORT", "4006"))

TOPICS = [
    "purchase.created",
    "purchase.voting.started",
    "purchase.voting.closed",
    "purchase.voting.tie",
    "purchase.vote.cast",
    "purchase.vote.changed",
    "purchase.candidate.added",
    "purchase.cancelled",
    "payment.topup.completed",
    "payment.hold.created",
    "payment.committed",
    "payment.released",
    "commission.held",
    "commission.committed",
    "commission.released",
    "escrow.created",
    "escrow.deposited",
    "escrow.confirmed",
    "escrow.released",
    "escrow.disputed",
    "review.created",
    "complaint.filed",
    "complaint.resolved",
    "user.auto_blocked",
    "search.query",
]

# ─── In-Memory Event Store (would be ClickHouse in production) ────────────────

event_store: list[dict[str, Any]] = []
purchase_stats: dict[str, dict] = {}   # purchaseId -> stats
payment_stats: dict[str, dict] = {}    # walletId -> stats
commission_stats: dict[str, dict] = {} # purchaseId -> commission stats
escrow_stats: dict[str, dict] = {}     # purchaseId -> escrow stats
reputation_stats: dict[str, dict] = {} # userId -> reputation stats
search_stats: dict[str, Any] = {"total_queries": 0, "avg_latency_ms": 0, "queries": []}


# ─── S3 Client ────────────────────────────────────────────────────────────────

def get_s3_client():
    return boto3.client(
        "s3",
        endpoint_url=S3_ENDPOINT,
        aws_access_key_id=S3_ACCESS_KEY,
        aws_secret_access_key=S3_SECRET_KEY,
        config=Config(signature_version="s3v4"),
        region_name="us-east-1",
    )


def ensure_bucket(s3) -> None:
    try:
        s3.head_bucket(Bucket=S3_BUCKET)
    except Exception:
        try:
            s3.create_bucket(Bucket=S3_BUCKET)
            logger.info(f"Created S3 bucket: {S3_BUCKET}")
        except Exception as e:
            logger.warning(f"Could not create bucket: {e}")


def upload_to_s3(key: str, data: bytes, content_type: str) -> str:
    try:
        s3 = get_s3_client()
        ensure_bucket(s3)
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=key,
            Body=data,
            ContentType=content_type,
        )
        url = f"{S3_ENDPOINT}/{S3_BUCKET}/{key}"
        logger.info(f"Uploaded to S3: {url}")
        return url
    except Exception as e:
        logger.error(f"S3 upload failed: {e}")
        raise


# ─── Report Generation ────────────────────────────────────────────────────────

def generate_purchases_xlsx() -> bytes:
    """Generate an XLSX summary of all purchase events."""
    rows = [e for e in event_store if "purchase" in e.get("topic", "")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Events"

    headers = ["Timestamp", "Topic", "Purchase ID", "Session ID", "Winner ID", "Total Votes", "User ID"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    for row in rows:
        payload = row.get("payload", {})
        ws.append([
            row.get("received_at", ""),
            row.get("topic", ""),
            payload.get("purchaseId", ""),
            payload.get("sessionId", ""),
            payload.get("winnerId", ""),
            payload.get("totalVotes", ""),
            payload.get("userId", payload.get("organizerId", "")),
        ])

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_payments_csv() -> bytes:
    """Generate a CSV of payment events."""
    rows = [e for e in event_store if "payment" in e.get("topic", "")]

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Timestamp", "Topic", "User ID", "Wallet ID", "Amount", "Currency", "Transaction ID", "Purchase ID"])
    for row in rows:
        p = row.get("payload", {})
        writer.writerow([
            row.get("received_at", ""),
            row.get("topic", ""),
            p.get("userId", ""),
            p.get("walletId", ""),
            p.get("amount", ""),
            p.get("currency", "RUB"),
            p.get("transactionId", ""),
            p.get("purchaseId", ""),
        ])
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def generate_vote_summary_xlsx() -> bytes:
    """Generate voting summary table."""
    votes = [e for e in event_store if e.get("topic") in (
        "purchase.vote.cast", "purchase.vote.changed", "purchase.voting.closed"
    )]

    df = pd.DataFrame([
        {
            "topic": e["topic"],
            "session_id": e["payload"].get("sessionId", ""),
            "purchase_id": e["payload"].get("purchaseId", ""),
            "user_id": e["payload"].get("userId", ""),
            "candidate_id": e["payload"].get("candidateId", e["payload"].get("newCandidateId", "")),
            "winner_id": e["payload"].get("winnerId", ""),
            "total_votes": e["payload"].get("totalVotes", 0),
            "ts": e["received_at"],
        }
        for e in votes
    ])

    buf = io.BytesIO()
    if not df.empty:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Votes", index=False)
            # Pivot: votes per candidate per session
            cast_df = df[df["topic"] == "purchase.vote.cast"]
            if not cast_df.empty:
                pivot = cast_df.groupby(["session_id", "candidate_id"]).size().reset_index(name="vote_count")
                pivot.to_excel(writer, sheet_name="Vote Tally", index=False)
    else:
        # Empty workbook
        wb = openpyxl.Workbook()
        wb.active.title = "Votes"
        wb.active.append(["No data yet"])
        wb.save(buf)

    return buf.getvalue()


# ─── Kafka Consumer ───────────────────────────────────────────────────────────

consumer_task: asyncio.Task | None = None


async def process_event(topic: str, payload: dict) -> None:
    """Process a single Kafka event and update in-memory stats."""
    received_at = datetime.now(timezone.utc).isoformat()
    event_store.append({
        "topic": topic,
        "payload": payload,
        "received_at": received_at,
    })

    # Update purchase stats
    if "purchaseId" in payload:
        pid = payload["purchaseId"]
        if pid not in purchase_stats:
            purchase_stats[pid] = {"events": 0, "votes": 0, "status": "unknown"}
        purchase_stats[pid]["events"] += 1
        if topic == "purchase.vote.cast":
            purchase_stats[pid]["votes"] += 1
        if topic == "purchase.voting.closed":
            purchase_stats[pid]["winner"] = payload.get("winnerId")
            purchase_stats[pid]["total_votes"] = payload.get("totalVotes", 0)

    # Update payment stats
    if "walletId" in payload or "userId" in payload:
        uid = payload.get("userId") or payload.get("walletId")
        if uid not in payment_stats:
            payment_stats[uid] = {"total_held": 0, "total_committed": 0, "total_released": 0}
        amount = payload.get("amount", 0)
        if topic == "payment.hold.created":
            payment_stats[uid]["total_held"] += amount
        elif topic == "payment.committed":
            payment_stats[uid]["total_committed"] += amount
        elif topic == "payment.released":
            payment_stats[uid]["total_released"] += amount

    # Update commission stats
    if topic.startswith("commission."):
        pid = payload.get("purchaseId", "unknown")
        if pid not in commission_stats:
            commission_stats[pid] = {"held": 0, "committed": 0, "released": 0, "percent": 0}
        amount = payload.get("amount", 0)
        if topic == "commission.held":
            commission_stats[pid]["held"] += amount
            commission_stats[pid]["percent"] = payload.get("percent", 0)
        elif topic == "commission.committed":
            commission_stats[pid]["committed"] += amount
        elif topic == "commission.released":
            commission_stats[pid]["released"] += amount

    # Update escrow stats
    if topic.startswith("escrow."):
        pid = payload.get("purchaseId", "unknown")
        if pid not in escrow_stats:
            escrow_stats[pid] = {"total_deposited": 0, "confirmations": 0, "required": 0, "status": "active"}
        if topic == "escrow.deposited":
            escrow_stats[pid]["total_deposited"] += payload.get("amount", 0)
        elif topic == "escrow.confirmed":
            escrow_stats[pid]["confirmations"] = payload.get("confirmationsReceived", 0)
            escrow_stats[pid]["required"] = payload.get("confirmationsRequired", 0)
        elif topic == "escrow.released":
            escrow_stats[pid]["status"] = "released"
        elif topic == "escrow.disputed":
            escrow_stats[pid]["status"] = "disputed"

    # Update reputation stats
    if topic in ("review.created", "complaint.filed", "complaint.resolved", "user.auto_blocked"):
        target_id = payload.get("targetId") or payload.get("userId", "unknown")
        if target_id not in reputation_stats:
            reputation_stats[target_id] = {"reviews": 0, "avg_rating": 0, "complaints": 0, "blocked": False}
        if topic == "review.created":
            stats = reputation_stats[target_id]
            stats["reviews"] += 1
            rating = payload.get("rating", 0)
            # Running average
            stats["avg_rating"] = ((stats["avg_rating"] * (stats["reviews"] - 1)) + rating) / stats["reviews"]
        elif topic == "complaint.filed":
            reputation_stats[target_id]["complaints"] += 1
        elif topic == "user.auto_blocked":
            reputation_stats[target_id]["blocked"] = True

    # Update search stats
    if topic == "search.query":
        search_stats["total_queries"] += 1

    # Periodically generate and upload reports (every 100 events)
    if len(event_store) % 100 == 0:
        await generate_and_upload_reports()


async def generate_and_upload_reports() -> None:
    """Generate all reports and upload to S3."""
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    loop = asyncio.get_event_loop()

    # Generate reports in executor (CPU-bound)
    try:
        xlsx_data = await loop.run_in_executor(None, generate_purchases_xlsx)
        upload_to_s3(f"reports/purchases_{ts}.xlsx", xlsx_data,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        logger.error(f"Failed to generate purchases report: {e}")

    try:
        csv_data = await loop.run_in_executor(None, generate_payments_csv)
        upload_to_s3(f"reports/payments_{ts}.csv", csv_data, "text/csv")
    except Exception as e:
        logger.error(f"Failed to generate payments report: {e}")

    try:
        vote_data = await loop.run_in_executor(None, generate_vote_summary_xlsx)
        upload_to_s3(f"reports/votes_{ts}.xlsx", vote_data,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        logger.error(f"Failed to generate votes report: {e}")


async def kafka_consumer_loop() -> None:
    consumer = AIOKafkaConsumer(
        *TOPICS,
        bootstrap_servers=KAFKA_BROKERS,
        group_id=KAFKA_GROUP_ID,
        auto_offset_reset="latest",
        enable_auto_commit=True,
        value_deserializer=lambda m: json.loads(m.decode("utf-8")),
    )

    retry_delay = 5
    while True:
        try:
            await consumer.start()
            logger.info(f"Kafka consumer started, topics: {TOPICS}")
            async for msg in consumer:
                try:
                    await process_event(msg.topic, msg.value)
                except Exception as e:
                    logger.error(f"Error processing {msg.topic}: {e}")
        except Exception as e:
            logger.error(f"Kafka consumer error: {e}. Retrying in {retry_delay}s...")
            await asyncio.sleep(retry_delay)
            retry_delay = min(retry_delay * 2, 60)
        finally:
            try:
                await consumer.stop()
            except Exception:
                pass


# ─── FastAPI App ──────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    global consumer_task
    consumer_task = asyncio.create_task(kafka_consumer_loop())
    logger.info("Analytics service started")
    yield
    if consumer_task:
        consumer_task.cancel()
        try:
            await consumer_task
        except asyncio.CancelledError:
            pass
    logger.info("Analytics service stopped")


app = FastAPI(
    title="Analytics Service",
    description="Kafka consumer that generates purchase/payment analytics reports",
    version="1.0.0",
    lifespan=lifespan,
)


@app.get("/health")
async def health():
    return {"status": "ok", "service": "analytics-service", "events_processed": len(event_store)}


@app.get("/stats/purchases")
async def get_purchase_stats():
    return {"success": True, "data": purchase_stats}


@app.get("/stats/payments")
async def get_payment_stats():
    return {"success": True, "data": payment_stats}


@app.get("/stats/commissions")
async def get_commission_stats():
    return {"success": True, "data": commission_stats}


@app.get("/stats/escrow")
async def get_escrow_stats():
    return {"success": True, "data": escrow_stats}


@app.get("/stats/reputation")
async def get_reputation_stats():
    return {"success": True, "data": reputation_stats}


@app.get("/stats/search")
async def get_search_stats():
    return {"success": True, "data": search_stats}


@app.get("/stats/summary")
async def get_summary():
    return {
        "success": True,
        "data": {
            "total_events": len(event_store),
            "purchases_tracked": len(purchase_stats),
            "users_tracked": len(payment_stats),
            "commissions_tracked": len(commission_stats),
            "escrow_accounts_tracked": len(escrow_stats),
            "reputation_profiles_tracked": len(reputation_stats),
            "search_queries": search_stats["total_queries"],
            "topics_consumed": TOPICS,
        },
    }


@app.post("/reports/generate")
async def trigger_report_generation():
    """Manually trigger report generation and S3 upload."""
    try:
        await generate_and_upload_reports()
        return {"success": True, "message": "Reports generated and uploaded to S3"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/purchases/download")
async def download_purchases_xlsx():
    """Generate and return purchases XLSX directly."""
    from fastapi.responses import Response
    data = generate_purchases_xlsx()
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=purchases.xlsx"},
    )


@app.get("/reports/payments/download")
async def download_payments_csv():
    """Generate and return payments CSV directly."""
    from fastapi.responses import Response
    data = generate_payments_csv()
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=payments.csv"},
    )


@app.get("/reports/votes/download")
async def download_votes_xlsx():
    """Generate and return vote summary XLSX directly."""
    from fastapi.responses import Response
    data = generate_vote_summary_xlsx()
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=votes.xlsx"},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=PORT, reload=False)
